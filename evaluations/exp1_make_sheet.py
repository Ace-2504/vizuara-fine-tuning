"""Exp 1 — build the blind human-grading workbook + hidden answer key.

Stratified sample of 10 answers per model x 15 = 150, balanced across case-law/sec/fineweb-edu,
with the AI judge's score and the model name HIDDEN. Writes:
  eval_results/exp1_grading_sheet.xlsx  -> Instructions + Grading (blank human fields, dropdowns)
  eval_results/exp1_key.json            -> row_id -> {model, pair_id, source, gemini parts/score,
                                                      luna score}  (NOT shown to grader; for kappa)
Human grades blind; judge_calibration.py later joins on row_id and computes weighted Cohen's kappa,
correlation, and % within 1 point. Nothing published.
"""
from __future__ import annotations
import json, io, os, random, re
from collections import defaultdict, Counter


def gradeable(s):
    """A human can actually read/grade this answer — excludes empty, control-only, and the
    degenerate char/word repetition loops the tiny base models fall into (125M/500M base are
    ~97% degenerate). Keeps real answers even if short or wrong."""
    c = "".join(ch for ch in (s or "") if ord(ch) >= 32 or ch in "\t\n").strip()
    if len(c) < 2:
        return False
    if max(Counter(c).values()) / len(c) > 0.5:            # single char dominates -> char loop
        return False
    words = re.findall(r"[A-Za-z]{2,}", c)
    if len(set(w.lower() for w in words)) >= 3:            # real vocabulary -> readable (even if repetitive)
        return True
    if len(c) <= 50 and re.search(r"[A-Za-z0-9]", c):      # legit short answer (e.g. "PURPA.", "August 1993")
        return True
    return False                                            # long but <3 distinct words = pure loop / junk

HERE = os.path.dirname(os.path.abspath(__file__)); ROOT = os.path.dirname(HERE)
ER = os.path.join(ROOT, "eval_results")
PER_MODEL = 10
SEED = 1337

MODELS = [
    ("125m-base","base-125m.judged.json"),("125m-qa","slm-125m-sft.judged.json"),
    ("125m-raft","slm-125m-raft.judged.json"),("125m-dpo","slm-125m-sft-dpo.judged.json"),
    ("125m-rlaif","slm-125m-sft-rlaif.judged.json"),("500m-base","base-500m.judged.json"),
    ("500m-qa","slm-500m-sft.judged.json"),("500m-raft","slm-500m-raft.judged.json"),
    ("500m-dpo","slm-500m-sft-dpo.judged.json"),("500m-rlaif","slm-500m-sft-rlaif.judged.json"),
    ("gemma-base","base-gemma.judged.json"),("gemma-qa","gemma-2-2b-sft.judged.json"),
    ("gemma-raft","gemma-2-2b-raft.judged.json"),("gemma-dpo","gemma-2-2b-sft-dpo.judged.json"),
    ("gemma-rlaif","gemma-2-2b-sft-rlaif.judged.json"),
]


def split_qc(user):
    i = user.rfind("Question:")
    q = user[i + len("Question:"):].strip() if i >= 0 else user
    ctx = user[:i].strip() if i >= 0 else ""
    return q, ctx


def stratified(items, k, rng):
    """Pick k items balanced across sources."""
    by = defaultdict(list)
    for it in items:
        by[it.get("source")].append(it)
    for s in by:
        rng.shuffle(by[s])
    srcs = sorted(by)
    out, i = [], 0
    while len(out) < k and any(by[s] for s in srcs):
        s = srcs[i % len(srcs)]
        if by[s]:
            out.append(by[s].pop())
        i += 1
    return out[:k]


def main():
    rng = random.Random(SEED)
    # luna scores (optional 3-way key)
    luna = {}
    lp = os.path.join(ER, "exp2_luna_scores.jsonl")
    if os.path.exists(lp):
        for line in io.open(lp, encoding="utf-8"):
            line = line.strip()
            if line:
                r = json.loads(line); luna[(r["model"], r["pair_id"])] = r["score"]

    rows = []
    for mid, fn in MODELS:
        p = os.path.join(ER, fn)
        items = [it for it in json.load(io.open(p, encoding="utf-8"))["per_item"]
                 if it["cond"] == "clean" and gradeable(it["resp"])]
        for it in stratified(items, PER_MODEL, rng):
            q, ctx = split_qc(it["user"])
            g = it.get("judge") or {}
            rows.append({
                "model": mid, "pair_id": it["pair_id"], "source": it.get("source"),
                "context": ctx, "question": q, "reference": it["ref"], "answer": it["resp"],
                "gemini_parts": g.get("parts"), "gemini_score": g.get("score"),
                "gemini_grounded": g.get("grounded"),
                "luna_score": luna.get((mid, it["pair_id"])),
            })
    rng.shuffle(rows)                                  # blind: break model grouping
    for i, r in enumerate(rows, 1):
        r["row_id"] = f"R{i:03d}"

    # ---- key (hidden) ----
    key = {r["row_id"]: {k: r[k] for k in ("model","pair_id","source","gemini_parts",
                                           "gemini_score","gemini_grounded","luna_score")} for r in rows}
    json.dump(key, io.open(os.path.join(ER, "exp1_key.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)

    # ---- workbook ----
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    def clean(v):
        return ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v
    wb = openpyxl.Workbook()

    # Instructions sheet
    ins = wb.active; ins.title = "Instructions"
    lines = [
        ("Exp 1 — Blind human grading", True),
        ("", False),
        ("You are grading answers from an unnamed QA system. The model identity and the AI judge's", False),
        ("score are hidden on purpose — grade only what you see. Fill the 5 grey columns per row.", False),
        ("", False),
        ("Score each answer against the RUBRIC (same one the AI judge used):", True),
        ("  correctness (0-5): factual agreement with the REFERENCE. 5 = fully right, 0 = wrong.", False),
        ("  completeness (0-2): covers the key points, not just one of them.", False),
        ("  groundedness (0-2): 2 = invents nothing; 0 = fabricated figures, cases or citations (judge vs CONTEXT).", False),
        ("  clarity (0-1): answers what was actually asked, without padding or contradiction.", False),
        ("  fabricated (y/n): does the answer state any number/fact/citation NOT supported by the context?", False),
        ("", False),
        ("Notes:", True),
        ("  - correctness+completeness+groundedness+clarity add to a total out of 10 (computed for you).", False),
        ("  - Judge meaning, not wording — a correct paraphrase scores high.", False),
        ("  - Base models often ramble/complete rather than answer — grade them on the same rubric.", False),
        ("  - Optional: have a friend grade ~40 rows in a copy for a human-ceiling estimate.", False),
        ("  - 150 rows, ~30-40 min if brisk. Save and send back the file when done.", False),
    ]
    for i, (txt, bold) in enumerate(lines, 1):
        c = ins.cell(row=i, column=1, value=txt); c.font = Font(bold=bold, size=13 if (bold and i == 1) else 11)
    ins.column_dimensions["A"].width = 110

    # Grading sheet
    ws = wb.create_sheet("Grading")
    headers = ["row_id","source","context","question","reference","answer",
               "correctness (0-5)","completeness (0-2)","groundedness (0-2)","clarity (0-1)",
               "fabricated (y/n)","total (auto)"]
    hdr_fill = PatternFill("solid", fgColor="305496"); grey = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="BBBBBB"); border = Border(*(thin,)*4)
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center"); c.border = border
    ws.freeze_panes = "A2"
    widths = [8,11,60,40,34,44,12,12,12,10,12,11]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

    for i, r in enumerate(rows, start=2):
        vals = [r["row_id"], r["source"], clean(r["context"]), clean(r["question"]),
                clean(r["reference"]), clean(r["answer"])]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.alignment = Alignment(wrap_text=True, vertical="top"); c.border = border
        for j in range(7, 12):                          # blank grey grade columns
            c = ws.cell(row=i, column=j); c.fill = grey; c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")
        # auto total
        L = openpyxl.utils.get_column_letter
        ws.cell(row=i, column=12,
                value=f"=IF(COUNT({L(7)}{i}:{L(10)}{i})=4,SUM({L(7)}{i}:{L(10)}{i}),\"\")").border = border
        ws.row_dimensions[i].height = 90

    # dropdown validations
    dv = {7: "0,1,2,3,4,5", 8: "0,1,2", 9: "0,1,2", 10: "0,1", 11: '"y,n"'}
    for col, opts in dv.items():
        formula = opts if opts.startswith('"') else f'"{opts}"'
        d = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(d)
        L = openpyxl.utils.get_column_letter(col)
        d.add(f"{L}2:{L}{len(rows)+1}")

    out = os.path.join(ER, "exp1_grading_sheet.xlsx")
    wb.save(out)
    # report
    from collections import Counter
    print(f"rows={len(rows)}  models={len(set(r['model'] for r in rows))}")
    print("per-source:", dict(Counter(r["source"] for r in rows)))
    print("per-model:", dict(Counter(r["model"] for r in rows)))
    print("wrote", out)
    print("wrote", os.path.join(ER, "exp1_key.json"), "(hidden key — do NOT open before grading)")


if __name__ == "__main__":
    main()
