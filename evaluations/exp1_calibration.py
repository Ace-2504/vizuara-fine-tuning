"""Exp 1 analysis — judge calibration against human grades (xlsx-sheet flow).

Run AFTER you fill in eval_results/exp1_grading_sheet.xlsx and save it:
    python evaluations/exp1_calibration.py --graded eval_results/exp1_grading_sheet.xlsx

Joins your grades to exp1_key.json on row_id and reports, for human-vs-Gemini (and human-vs-luna
if present): quadratic-weighted Cohen's kappa on the 0-10 total, Pearson/Spearman correlation,
% of answers within 1 point, per-dimension kappa, and fabricated-flag agreement.
Targets: kappa >= 0.6, r >= 0.7. Output: eval_results/STATS_exp1_calibration.{md,json}. Unpublished.
"""
from __future__ import annotations
import argparse, json, io, os
import numpy as np
from scipy import stats

HERE = os.path.dirname(os.path.abspath(__file__)); ROOT = os.path.dirname(HERE)
ER = os.path.join(ROOT, "eval_results")


def quad_weighted_kappa(a, b, minv, maxv):
    a = np.asarray(a, float); b = np.asarray(b, float)
    scale = 2 if (np.any(a % 1 != 0) or np.any(b % 1 != 0)) else 1
    lo, hi = int(round(minv * scale)), int(round(maxv * scale))
    ai = np.rint(a * scale).astype(int) - lo
    bi = np.rint(b * scale).astype(int) - lo
    N = hi - lo + 1
    if N <= 1: return float("nan")
    O = np.zeros((N, N))
    for x, y in zip(ai, bi): O[x, y] += 1
    w = np.fromfunction(lambda i, j: (i - j) ** 2 / (N - 1) ** 2, (N, N))
    act = O.sum(1); pred = O.sum(0); n = O.sum()
    E = np.outer(act, pred) / n
    den = (w * E).sum()
    return 1 - (w * O).sum() / den if den else float("nan")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--graded", required=True)
    a = ap.parse_args()
    import openpyxl
    key = json.load(io.open(os.path.join(ER, "exp1_key.json"), encoding="utf-8"))
    ws = openpyxl.load_workbook(a.graded, data_only=True)["Grading"]
    hdr = [str(c.value) for c in ws[1]]
    col = {h.split(" ")[0]: i for i, h in enumerate(hdr)}

    H_tot, G_tot, L_tot, H_fab, G_fab = [], [], [], [], []
    per_dim = {d: ([], []) for d in ("correctness", "completeness", "groundedness", "clarity")}
    ungraded = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        rid = r[col["row_id"]]
        if rid not in key: continue
        vals, ok = {}, True
        for d in per_dim:
            v = r[col[d]]
            if v is None or v == "": ok = False; break
            vals[d] = float(v)
        if not ok: ungraded += 1; continue
        k = key[rid]
        if k.get("gemini_score") is None: continue
        htot = sum(vals.values()); gp = k.get("gemini_parts") or {}
        H_tot.append(htot); G_tot.append(float(k["gemini_score"]))
        if k.get("luna_score") is not None: L_tot.append((htot, float(k["luna_score"])))
        for d in per_dim:
            if gp.get(d) is not None:
                per_dim[d][0].append(vals[d]); per_dim[d][1].append(float(gp[d]))
        fab = r[col["fabricated"]]
        if fab is not None and str(fab).strip():
            H_fab.append(1 if str(fab).strip().lower().startswith("y") else 0)
            G_fab.append(0 if k.get("gemini_grounded") else 1)

    n = len(H_tot)
    if n == 0: raise SystemExit("no graded rows — fill the grey columns and save first.")
    H, G = np.array(H_tot), np.array(G_tot)
    kappa = quad_weighted_kappa(H, G, 0, 10)
    pear = stats.pearsonr(H, G); sp = stats.spearmanr(H, G)
    within1 = float(np.mean(np.abs(H - G) <= 1.0) * 100); mae = float(np.abs(H - G).mean())
    md = ["# Exp 1 — Judge calibration (human vs Gemini primary judge)\n",
          f"Graded rows: **{n}** (ungraded/skipped: {ungraded}).\n",
          "## Human vs Gemini — total (0-10)\n",
          f"- **Quadratic-weighted Cohen's kappa = {kappa:.3f}** (target >= 0.60: {'PASS' if kappa>=0.6 else 'BELOW'})",
          f"- **Pearson r = {pear[0]:.3f}** (p={pear[1]:.1e}; target >= 0.70: {'PASS' if pear[0]>=0.7 else 'BELOW'})",
          f"- Spearman rho = {sp.correlation:.3f}   |   Within 1 point: **{within1:.1f}%**   |   MAE {mae:.2f}",
          "", "## Per-dimension weighted kappa\n", "| Dimension | n | kappa |", "|---|---|---|"]
    out = {"n": n, "kappa_total": kappa, "pearson": float(pear[0]), "spearman": float(sp.correlation),
           "within1_pct": within1, "mae": mae, "per_dim": {}}
    dim_max = {"correctness": 5, "completeness": 2, "groundedness": 2, "clarity": 1}
    for d in ("correctness", "completeness", "groundedness", "clarity"):
        hh, gg = per_dim[d]
        if hh:
            kd = quad_weighted_kappa(hh, gg, 0, dim_max[d])
            md.append(f"| {d} | {len(hh)} | {kd:.3f} |"); out["per_dim"][d] = kd
    if H_fab:
        hf, gf = np.array(H_fab), np.array(G_fab)
        po = (hf == gf).mean()
        pe = hf.mean()*gf.mean() + (1-hf.mean())*(1-gf.mean())
        kfab = (po - pe)/(1 - pe) if pe < 1 else float("nan")
        md += ["", f"## Fabrication flag (human y/n vs judge not-grounded)",
               f"- Agreement {po*100:.1f}%, Cohen's kappa {kfab:.3f} (n={len(hf)})"]
        out["fab_agree_pct"] = float(po*100); out["fab_kappa"] = float(kfab)
    if L_tot:
        Lh = np.array([x for x,_ in L_tot]); Ll = np.array([y for _,y in L_tot])
        md += ["", "## Bonus — human vs luna",
               f"- weighted kappa {quad_weighted_kappa(Lh,Ll,0,10):.3f}, Pearson r {stats.pearsonr(Lh,Ll)[0]:.3f}, "
               f"within 1pt {np.mean(np.abs(Lh-Ll)<=1)*100:.1f}%"]
    md += ["", "## Verdict",
           (f"Human vs judge: kappa={kappa:.2f}, r={pear[0]:.2f}, {within1:.0f}% within 1 point — "
            + ("judge is well-calibrated to human grading."
               if (kappa>=0.6 and pear[0]>=0.7) else "moderate calibration; report figures, treat scores as ordinal."))]
    io.open(os.path.join(ER, "STATS_exp1_calibration.md"), "w", encoding="utf-8").write("\n".join(md)+"\n")
    json.dump(out, io.open(os.path.join(ER, "STATS_exp1_calibration.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)
    print("\n".join(md))


if __name__ == "__main__":
    main()
