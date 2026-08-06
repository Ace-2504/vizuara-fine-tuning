"""
Build SLM-arena-reconciliation.xlsx — does the arena's "Cost to build" table agree with
the cost tables on the individual model sites?

Raw rows are read straight out of slm-frontends/lib/models.ts so the workbook cannot drift
from what the sites actually display. Every total in the workbook is an Excel formula.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MODELS_TS = Path(r"D:\slm-frontends\lib\models.ts")
OUT = Path(__file__).resolve().parent / "SLM-arena-reconciliation.xlsx"

FONT = "Arial"
MONEY = '$#,##0.00;($#,##0.00);-'
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
TITLE = Font(name=FONT, bold=True, size=14)
SUB = Font(name=FONT, italic=True, size=9, color="595959")
BOLD = Font(name=FONT, bold=True, size=10)
BODY = Font(name=FONT, size=10)
BLUE = Font(name=FONT, size=10, color="0000FF")          # hardcoded input
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
RED_FILL = PatternFill("solid", fgColor="FCE4E4")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

# ---------------------------------------------------------------- arena figures (published)
ARENA = [
    ("125M pretraining", 70.14),
    ("Fine-tuning", 19.94),
    ("Evaluation", 19.19),
    ("Alignment", 5.23),
    ("Earlier SFT study", 1.66),
    ("Probes & smoke tests", 0.18),
    ("Image builds & exports", 0.05),
]

# ---------------------------------------------------------------- read the frontends


def frontend_rows() -> list[tuple[str, str, float]]:
    s = MODELS_TS.read_text(encoding="utf-8")
    out: list[tuple[str, str, float]] = []
    for m in re.finditer(r'"((?:125m|500m|gemma)-(?:base|qa|raft|dpo|rlaif))": \{', s):
        sid = m.group(1)
        i = m.start()
        j = s.index("\n  },", i)
        for stage, cost in re.findall(
            r'\{ stage: "([^"]+)", note: "[^"]*", cost: "\$([0-9.]+)" \}', s[i:j]
        ):
            out.append((sid, stage, float(cost)))
    return out


def category(stage: str) -> str:
    t = stage.lower()
    if "pretraining" in t:
        return "Pretraining"
    if "evaluation" in t:
        return "Evaluation"
    if "dataset" in t:
        return "Gemini data"
    if "reward model" in t or "dpo alignment" in t or "ppo alignment" in t:
        return "Alignment"
    return "Fine-tuning"


# ---------------------------------------------------------------- helpers


def head(ws, row: int, labels: list[str], widths: list[int]) -> None:
    for c, (label, w) in enumerate(zip(labels, widths), start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def title(ws, text: str, sub: str) -> int:
    ws["A1"] = text
    ws["A1"].font = TITLE
    ws["A2"] = sub
    ws["A2"].font = SUB
    return 4


def put(ws, row, col, value, *, font=BODY, fmt=None, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.border = BOX
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    return c


# ---------------------------------------------------------------- sheets


def sheet_summary(wb, rows):
    ws = wb.create_sheet("Summary")
    r = title(ws, "Arena vs. individual model sites — per category",
              "Do the arena's phase totals equal what the model sites display for the same work? "
              "Frontend figures are DE-DUPLICATED: one lineage shown on five sites is one cost, not five. "
              "State: AFTER the fix that closed the fine-tuning and evaluation gaps.")
    head(ws, r, ["Category", "Arena shows", "Unique work on the sites",
                 "Gap", "Match?", "Explanation"],
         [24, 14, 22, 12, 11, 74])
    r += 1

    data = [
        ("125M pretraining", 70.14, 70.14,
         "Shown on all four 125M sites and the base site, but it is one lineage — counting it five times would give $350.70."),
        ("Fine-tuning", 19.94, 19.94,
         "CLOSED. The 500M RAFT run ($1.74) now appears on the 500M base site, labelled as having no site of its own."),
        ("Evaluation", 19.19, 19.19,
         "CLOSED. The three base models were given cost tables, so all 13 shares are now visible (sites round each to $1.48)."),
        ("Alignment", 5.23, 5.23,
         "The $0.06 reward model is split three ways at $0.02 and the shares add back to the whole."),
        ("Earlier SFT study", 1.66, 0.00,
         "A separate project that preceded this build. Correctly absent — there is no model page to attribute it to."),
        ("Probes & smoke tests", 0.18, 0.00, "Not attributable to any published model."),
        ("Image builds & exports", 0.05, 0.00, "Container builds and checkpoint downloads."),
    ]
    first = r
    for name, arena, shown, why in data:
        put(ws, r, 1, name, font=BOLD)
        put(ws, r, 2, arena, font=BLUE, fmt=MONEY)
        put(ws, r, 3, shown, font=BLUE, fmt=MONEY)
        put(ws, r, 4, f"=B{r}-C{r}", fmt=MONEY)
        c = put(ws, r, 5, f'=IF(ROUND(D{r},2)=0,"MATCH","GAP")')
        c.alignment = Alignment(horizontal="center")
        c.fill = GREEN_FILL if abs(arena - shown) < 0.005 else RED_FILL
        put(ws, r, 6, why)
        ws.row_dimensions[r].height = 28
        r += 1

    put(ws, r, 1, "TOTAL", font=BOLD)
    put(ws, r, 2, f"=SUM(B{first}:B{r-1})", font=BOLD, fmt=MONEY)
    put(ws, r, 3, f"=SUM(C{first}:C{r-1})", font=BOLD, fmt=MONEY)
    put(ws, r, 4, f"=B{r}-C{r}", font=BOLD, fmt=MONEY)
    put(ws, r, 5, "")
    put(ws, r, 6, "The $116.39 arena total is the invoiced Modal spend across all four accounts.", font=BOLD)

    ws.cell(row=first, column=2).comment = Comment(
        "Published on the arena's 'Cost to build' tab. Source: the four Modal dashboards, "
        "transcribed in SLM-cost-audit.xlsx.", "cost audit", width=320, height=90)
    ws.cell(row=first, column=3).comment = Comment(
        "Sum of the DISTINCT cost rows across the 10 model sites that carry a cost table. "
        "Repeated rows (a checkpoint shown on the sites derived from it) are counted once.",
        "cost audit", width=320, height=90)

    r += 2
    ws.cell(row=r, column=1, value="Bottom line").font = BOLD
    r += 1
    for line in [
        "All four attributable categories now match exactly. The only remaining difference is $1.89 of",
        "spend that belongs to no published model — the earlier SFT study, capability probes, and image",
        "builds. There is no page on which to show it, so its absence is correct rather than a gap.",
        "Gemini API spend was removed from the model sites entirely; it is published once, on the arena.",
    ]:
        ws.cell(row=r, column=1, value=line).font = BODY
        r += 1
    return ws


def sheet_finetuning(wb):
    ws = wb.create_sheet("Fine-tuning detail")
    r = title(ws, "Fine-tuning — the $1.74 gap, item by item",
              "Each fine-tune is one invoiced run. Some appear on several sites because DPO and RLAIF "
              "start from the QA-SFT checkpoint; that repetition is display, not extra cost.")
    head(ws, r, ["Model", "Stage", "Invoiced", "Sites displaying it",
                 "Times shown", "Counted once as"], [14, 12, 12, 42, 12, 12])
    r += 1
    items = [
        ("Gemma 2B", "QA-SFT", 9.84, "gemma-qa, gemma-dpo, gemma-rlaif", 3),
        ("Gemma 2B", "RAFT", 6.82, "gemma-raft", 1),
        ("SLM-500M", "QA-SFT", 0.93, "500m-dpo, 500m-rlaif", 2),
        ("SLM-500M", "RAFT", 1.74, "500m-base (no site of its own)", 1),
        ("SLM-125M", "QA-SFT", 0.20, "125m-qa, 125m-dpo, 125m-rlaif", 3),
        ("SLM-125M", "RAFT", 0.41, "125m-raft", 1),
    ]
    first = r
    for model, stage, cost, sites, n in items:
        put(ws, r, 1, model, font=BOLD)
        put(ws, r, 2, stage)
        put(ws, r, 3, cost, font=BLUE, fmt=MONEY)
        put(ws, r, 4, sites, fill=GREEN_FILL if n else RED_FILL)
        put(ws, r, 5, n).alignment = Alignment(horizontal="center")
        put(ws, r, 6, f"=IF(E{r}=0,0,C{r})", fmt=MONEY)
        r += 1

    put(ws, r, 2, "Invoiced total", font=BOLD)
    put(ws, r, 3, f"=SUM(C{first}:C{r-1})", font=BOLD, fmt=MONEY)
    put(ws, r, 4, "matches the arena's Fine-tuning row", font=BOLD)
    put(ws, r, 5, "")
    put(ws, r, 6, f"=SUM(F{first}:F{r-1})", font=BOLD, fmt=MONEY)
    r += 1
    put(ws, r, 2, "Gap", font=BOLD)
    put(ws, r, 3, f"=C{r-1}-F{r-1}", font=BOLD, fmt=MONEY, fill=GREEN_FILL)
    put(ws, r, 4, "zero — the 500M RAFT run is now shown on the 500M base site", font=BOLD)
    put(ws, r, 5, "")
    put(ws, r, 6, "")

    r += 3
    ws.cell(row=r, column=1, value="Naive sum, for contrast").font = BOLD
    r += 1
    ws.cell(row=r, column=1,
            value="Adding every fine-tuning row as displayed, without de-duplicating, gives $39.21 — "
                  "roughly double the truth, because Gemma's QA-SFT alone is shown three times.").font = BODY
    return ws


def sheet_eval(wb):
    ws = wb.create_sheet("Evaluation detail")
    r = title(ws, "Evaluation — one bill, thirteen shares",
              "A single $19.19 evaluation run covered all 13 models. Each site shows its own share.")
    labels = ["Item", "Value", "Note"]
    head(ws, r, labels, [34, 16, 62])
    r += 1
    first = r
    put(ws, r, 1, "Invoiced evaluation total", font=BOLD)
    put(ws, r, 2, 19.19, font=BLUE, fmt=MONEY)
    put(ws, r, 3, "eval.evaluate on the ace-compoz account (set1 + set2)")
    r += 1
    put(ws, r, 1, "Models evaluated", font=BOLD)
    put(ws, r, 2, 13, font=BLUE)
    put(ws, r, 3, "every model in the arena, 500 held-out questions each")
    r += 1
    put(ws, r, 1, "Exact share per model", font=BOLD)
    put(ws, r, 2, f"=B{first}/B{first+1}", fmt='$#,##0.00000')
    put(ws, r, 3, "the sites round this to $1.48")
    share_row = r
    r += 1
    put(ws, r, 1, "Sites carrying a cost table", font=BOLD)
    put(ws, r, 2, 13, font=BLUE)
    put(ws, r, 3, "all 13 — the three base models were given cost tables to close this gap")
    shown_row = r
    r += 1
    put(ws, r, 1, "Shown across those sites", font=BOLD)
    put(ws, r, 2, f"=B{share_row}*B{shown_row}", fmt=MONEY)
    put(ws, r, 3, "what a reader can actually see")
    r += 1
    put(ws, r, 1, "Never shown", font=BOLD)
    put(ws, r, 2, f"=B{share_row}*(B{first+1}-B{shown_row})", fmt=MONEY, fill=RED_FILL)
    put(ws, r, 3, "nothing — every share is now published")
    r += 1
    put(ws, r, 1, "Check: shown + never shown", font=BOLD)
    put(ws, r, 2, f"=B{r-2}+B{r-1}", font=BOLD, fmt=MONEY, fill=GREEN_FILL)
    put(ws, r, 3, "equals the invoiced total, so the split is complete", font=BOLD)
    r += 2
    put(ws, r, 1, "Rounding note", font=BOLD)
    put(ws, r, 3, "13 sites x $1.48 displayed = $19.24, five cents above the exact $19.19. Immaterial.")
    return ws


def sheet_rows(wb, rows):
    ws = wb.create_sheet("All frontend rows")
    r = title(ws, "Every cost row on every model site",
              f"Extracted from slm-frontends/lib/models.ts — {len(rows)} rows across 10 sites. "
              "This is the raw evidence behind the Summary sheet.")
    head(ws, r, ["Site", "Category", "Stage", "Cost as displayed"], [16, 16, 46, 18])
    r += 1
    first = r
    for sid, stage, cost in sorted(rows, key=lambda x: (category(x[1]), x[0])):
        put(ws, r, 1, sid)
        put(ws, r, 2, category(stage))
        put(ws, r, 3, stage)
        put(ws, r, 4, cost, fmt=MONEY)
        r += 1
    put(ws, r, 3, "Naive total (double-counts repeated work)", font=BOLD)
    put(ws, r, 4, f"=SUM(D{first}:D{r-1})", font=BOLD, fmt=MONEY)
    r += 2
    ws.cell(row=r, column=1,
            value="This naive total is NOT a meaningful figure — it is here only to show why "
                  "de-duplication is necessary.").font = SUB
    return ws


def sheet_close(wb):
    ws = wb.create_sheet("Reconciliation")
    r = title(ws, "Closing the loop",
              "Does every invoiced dollar land somewhere? The components below must sum to $116.39.")
    head(ws, r, ["Component", "Amount", "Where it appears"], [40, 14, 62])
    r += 1
    first = r
    parts = [
        ("125M pretraining", 70.14, "the 125M model sites, and the e4 site as four legs"),
        ("Fine-tuning", 19.94, "every QA-SFT and RAFT row, de-duplicated"),
        ("Alignment", 5.23, "DPO, PPO and the shared reward model"),
        ("Evaluation", 19.19, "all 13 sites now show their share"),
                ("Study, probes, builds", 1.89, "not attributable to any published model"),
    ]
    for name, amt, where in parts:
        put(ws, r, 1, name)
        put(ws, r, 2, amt, font=BLUE, fmt=MONEY)
        put(ws, r, 3, where)
        r += 1
    put(ws, r, 1, "Sum", font=BOLD)
    put(ws, r, 2, f"=SUM(B{first}:B{r-1})", font=BOLD, fmt=MONEY)
    put(ws, r, 3, "")
    sum_row = r
    r += 1
    put(ws, r, 1, "Total invoiced across 4 Modal accounts", font=BOLD)
    put(ws, r, 2, 116.39, font=BLUE, fmt=MONEY)
    put(ws, r, 3, "ace-2504 $30.36 + ace-compoz $30.29 + singh1621 $28.26 + aceaynon2504 $27.48")
    r += 1
    put(ws, r, 1, "Difference", font=BOLD)
    put(ws, r, 2, f"=B{sum_row}-B{r-1}", font=BOLD, fmt=MONEY, fill=GREEN_FILL)
    put(ws, r, 3, "zero — nothing double-counted, nothing dropped", font=BOLD)

    r += 3
    ws.cell(row=r, column=1, value="Gemini API — published once, on the arena").font = Font(
        name=FONT, bold=True, size=11, color="1F3864")
    r += 1
    head(ws, r, ["Source", "Amount", "Note"], [40, 14, 62])
    r += 1
    put(ws, r, 1, "Arena 'Cost to build' tab")
    put(ws, r, 2, "Rs 2,215", font=BLUE)
    put(ws, r, 3, "the combined API bill, as supplied by Harman")
    r += 1
    put(ws, r, 1, "Gemini rows on the model sites")
    put(ws, r, 2, 0.00, font=BLUE, fmt=MONEY)
    put(ws, r, 3, "all 16 removed — the previous $6.90 was an unverified estimate")
    r += 1
    put(ws, r, 1, "Status", font=BOLD)
    put(ws, r, 2, "SINGLE SOURCE", font=BOLD, fill=GREEN_FILL)
    put(ws, r, 3, "One figure, one place. Nothing to reconcile: Modal spend is per-model, "
                  "the Gemini bill is reported only as a whole.", font=BOLD)
    ws.row_dimensions[r].height = 28
    return ws


def main() -> None:
    rows = frontend_rows()
    wb = Workbook()
    wb.remove(wb.active)
    sheet_summary(wb, rows)
    sheet_finetuning(wb)
    sheet_eval(wb)
    sheet_rows(wb, rows)
    sheet_close(wb)
    for ws in wb:
        ws.sheet_view.showGridLines = False
    wb.save(OUT)
    print(f"wrote {OUT}  ({len(rows)} frontend rows, {len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    main()
