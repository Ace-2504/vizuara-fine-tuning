"""
Verify the formulas in SLM-arena-reconciliation.xlsx.

LibreOffice is not installed here, so the skill's recalc.py cannot run. This evaluates the
subset of formula shapes the workbook actually uses (SUM over a range, +, -, *, /, IF) against
the literal cell values, which is what catches the real risk: a range that points at the wrong
rows. Any formula shape it does not understand is reported rather than silently passed.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

WB = Path(__file__).resolve().parent / "SLM-arena-reconciliation.xlsx"

EXPECTED = {  # sheet -> {cell: expected value}, computed independently by hand
    "Summary": {"B12": 116.39, "C12": 114.50, "D12": 1.89,
                "D5": 0.0, "D6": 0.0, "D7": 0.0, "D8": 0.0},
    "Fine-tuning detail": {"C11": 19.94, "F11": 19.94, "C12": 0.0},
    "Evaluation detail": {"B7": 19.19 / 13, "B9": 19.19, "B10": 0.0, "B11": 19.19},
    "Reconciliation": {"B10": 116.39, "B12": 0.0},
}


def cellval(ws, ref: str):
    v = ws[ref].value
    return 0.0 if v is None else v


def evaluate(ws, formula: str) -> float | str:
    f = formula.lstrip("=")

    m = re.fullmatch(r'IF\(ROUND\((\w+),2\)=0,"MATCH","GAP"\)', f)
    if m:
        return "MATCH" if round(float(evaluate_ref(ws, m.group(1))), 2) == 0 else "GAP"
    m = re.fullmatch(r"IF\((\w+)=0,0,(\w+)\)", f)
    if m:
        return 0.0 if float(evaluate_ref(ws, m.group(1))) == 0 else float(evaluate_ref(ws, m.group(2)))
    m = re.fullmatch(r"SUM\((\w+):(\w+)\)", f)
    if m:
        lo, hi = m.group(1), m.group(2)
        c1, r1, c2, r2 = range_boundaries(f"{lo}:{hi}")
        tot = 0.0
        for row in range(r1, r2 + 1):
            for col in range(c1, c2 + 1):
                v = ws.cell(row=row, column=col).value
                if isinstance(v, (int, float)):
                    tot += v
                elif isinstance(v, str) and v.startswith("="):
                    got = evaluate(ws, v)
                    if isinstance(got, (int, float)):
                        tot += got
        return tot
    m = re.fullmatch(r"(\w+)\*\((\w+)-(\w+)\)", f)
    if m:
        a, b, c = (float(evaluate_ref(ws, g)) for g in m.groups())
        return a * (b - c)
    m = re.fullmatch(r"(\w+)([-+*/])(\w+)", f)
    if m:
        a, op, b = m.group(1), m.group(2), m.group(3)
        x, y = float(evaluate_ref(ws, a)), float(evaluate_ref(ws, b))
        if op == "-":
            return x - y
        if op == "+":
            return x + y
        if op == "*":
            return x * y
        return x / y if y else f"DIV0({a}/{b})"
    return f"UNPARSED({f})"


def evaluate_ref(ws, ref: str):
    v = cellval(ws, ref)
    if isinstance(v, str) and v.startswith("="):
        return evaluate(ws, v)
    return v


def main() -> int:
    wb = load_workbook(WB)
    total = bad = unparsed = 0
    print(f"checking {WB.name}\n" + "-" * 74)

    for ws in wb:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    total += 1
                    got = evaluate(ws, c.value)
                    if isinstance(got, str) and got.startswith("UNPARSED"):
                        unparsed += 1
                        print(f"  ?  {ws.title}!{c.coordinate}  {got}")

    for sheet, cells in EXPECTED.items():
        ws = wb[sheet]
        for ref, want in cells.items():
            got = evaluate(ws, ws[ref].value) if isinstance(ws[ref].value, str) else ws[ref].value
            ok = isinstance(got, (int, float)) and abs(got - want) < 0.005
            bad += not ok
            print(f"  {'OK  ' if ok else 'FAIL'} {sheet}!{ref:5s} expected {want:>10.4f}"
                  f"   got {got if not isinstance(got, (int, float)) else round(got, 4)}")

    print("-" * 74)
    print(f"{total} formulas, {unparsed} unrecognised, {bad} wrong")
    return 1 if (bad or unparsed) else 0


if __name__ == "__main__":
    raise SystemExit(main())
