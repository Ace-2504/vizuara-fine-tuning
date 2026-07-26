"""Reconcile invoiced Modal spend against the costs shown on the frontends.

Inputs
  cost-audit/modal_<workspace>.json   invoiced per-app spend, transcribed from the four dashboards
  (frontend claims are hard-coded below, read out of slm-frontends/lib/models.ts and the e4 site)

Outputs
  cost-audit/SLM-cost-audit.xlsx   one sheet per Modal account + summary + reconciliation
  cost-audit/COST-AUDIT.md         the written report

The frontends are NOT modified.
"""
from __future__ import annotations
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))

ACCOUNTS = ["ace-2504", "aceaynon2504", "singh1621", "ace-compoz"]

# Modal app -> (what it actually was, which frontend line claims it, claimed $ or None if never shown)
# "claimed" is the figure a site displays for that same work. None => the site never mentions it.
MAP = {
    ("ace-2504", "modal_train.pretrain"): (
        "125M pretraining — v1 (2.04B tokens, 1 epoch)", "e4 site · cost table · 'v1 pretraining'", 10.71),
    ("ace-2504", "train_gemma.train"): (
        "Gemma QLoRA fine-tune — earlier run on this account", "Gemma QA+RAFT sites · 'QLoRA fine-tune'", None),
    ("ace-2504", "modal_continue_train.continue_train"): (
        "125M pretraining — extension (+515M tokens)", "e4 site · cost table · 'Extension'", 2.70),
    ("ace-2504", "train_500m.train"): (
        "500M QA-SFT + RAFT fine-tune", "500M sites · 'QA supervised fine-tune' + RAFT", 1.23),
    ("ace-2504", "modal_sft.train"): (
        "125M SFT — earlier data-scaling study (separate project)", None, None),
    ("ace-2504", "modal_phase0_probe.probe"): (
        "Phase-0 capability probe", None, None),
    ("ace-2504", "<image build>"): ("Container image build", None, None),
    ("ace-2504", "modal_smoketest.check"): ("Smoke test", None, None),
    ("ace-2504", "export_sft_hf.export"): ("Export to HuggingFace", None, None),
    ("ace-2504", "modal_export_v1ext.export"): ("Export to HuggingFace", None, None),
    ("ace-2504", "modal_export_hf.export"): ("Export to HuggingFace", None, None),
    ("ace-2504", "modal_export_v1ext.verify"): ("Export verification", None, None),

    ("aceaynon2504", "modal_train_2epoch.train"): (
        "125M pretraining — e2 (2 epochs on the rebuilt 2.5B corpus)", "e4 site · cost table · 'e2'", 26.31),
    ("aceaynon2504", "<image build>"): ("Container image build", None, None),
    ("aceaynon2504", "modal_export_v1ext.export"): ("Export to HuggingFace", None, None),

    ("singh1621", "modal_train_e4.train"): (
        "125M pretraining — e4 (2 further epochs, 4 total)", "e4 site · cost table · 'e4 (this model)'", 18.06),
    ("singh1621", "<image build>"): ("Container image build", None, None),
    ("singh1621", "modal_export_v1ext.export"): ("Export to HuggingFace", None, None),

    ("ace-compoz", "eval.evaluate"): (
        "Evaluation — set1 + set2 across all 13 versions", None, None),
    ("ace-compoz", "train_gemma.train"): (
        "Gemma QLoRA fine-tune — QA-SFT + RAFT (final runs)", "Gemma QA+RAFT sites · 'QLoRA fine-tune'", 4.56),
    ("ace-compoz", "train_ppo.train_gemma"): (
        "Gemma RLAIF — PPO", "Gemma RLAIF site · 'PPO alignment'", 1.26),
    ("ace-compoz", "train_125m.train"): (
        "125M QA-SFT + RAFT fine-tune", "125M QA+RAFT sites · fine-tune rows", 0.39),
    ("ace-compoz", "train_dpo.train_gemma"): (
        "Gemma DPO", "Gemma DPO site · 'QLoRA-DPO alignment'", 0.09),
    ("ace-compoz", "train_ppo.train_500m"): (
        "500M RLAIF — PPO", "500M RLAIF site · 'PPO alignment'", 0.28),
    ("ace-compoz", "train_ppo.train_125m"): (
        "125M RLAIF — PPO", "125M RLAIF site · 'PPO alignment'", 0.07),
    ("ace-compoz", "train_dpo.train_500m"): (
        "500M DPO", "500M DPO site · 'DPO alignment'", 0.05),
    ("ace-compoz", "train_reward.train"): (
        "Bradley-Terry reward model (shared by all RLAIF runs)", "RLAIF sites · 'Reward model'", 0.03),
    ("ace-compoz", "train_dpo.train_125m"): (
        "125M DPO", "125M DPO site · 'DPO alignment'", 0.01),
    ("ace-compoz", "<image build>"): ("Container image build", None, None),
}

# Phase buckets for the spend report
PHASE = {
    "modal_train.pretrain": "125M pretraining",
    "modal_continue_train.continue_train": "125M pretraining",
    "modal_train_2epoch.train": "125M pretraining",
    "modal_train_e4.train": "125M pretraining",
    "train_gemma.train": "Fine-tuning (SFT/RAFT)",
    "train_500m.train": "Fine-tuning (SFT/RAFT)",
    "train_125m.train": "Fine-tuning (SFT/RAFT)",
    "modal_sft.train": "Earlier SFT study",
    "train_dpo.train_gemma": "Alignment (DPO/RLAIF)",
    "train_dpo.train_500m": "Alignment (DPO/RLAIF)",
    "train_dpo.train_125m": "Alignment (DPO/RLAIF)",
    "train_ppo.train_gemma": "Alignment (DPO/RLAIF)",
    "train_ppo.train_500m": "Alignment (DPO/RLAIF)",
    "train_ppo.train_125m": "Alignment (DPO/RLAIF)",
    "train_reward.train": "Alignment (DPO/RLAIF)",
    "eval.evaluate": "Evaluation",
    "modal_phase0_probe.probe": "Probes & tests",
    "modal_smoketest.check": "Probes & tests",
}


def load():
    out = {}
    for a in ACCOUNTS:
        p = os.path.join(HERE, f"modal_{a}.json")
        out[a] = json.load(open(p, encoding="utf-8"))
    return out


HDR = PatternFill("solid", fgColor="1F3864")
HDRF = Font(bold=True, color="FFFFFF", size=11)
BOLD = Font(bold=True)
OVER = PatternFill("solid", fgColor="FCE4E4")     # invoiced > claimed
NA = PatternFill("solid", fgColor="F2F2F2")
TOTROW = PatternFill("solid", fgColor="DDEBF7")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def style_header(ws, row=1):
    for c in ws[row]:
        if c.value is not None:
            c.fill, c.font, c.alignment = HDR, HDRF, Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def account_sheet(wb, acct, data):
    ws = wb.create_sheet(acct)
    ws.append(["Modal app", "What it actually was", "Invoiced (USD)",
               "Shown on frontend as", "Frontend claims (USD)", "Difference", "Status"])
    style_header(ws)
    inv_t = clm_t = 0.0
    for a in data["apps"]:
        app, usd = a["app"], a["usd"]
        what, where, claimed = MAP.get((acct, app), (app, None, None))
        inv_t += usd
        if claimed is None:
            row = [app, what, usd, where or "not shown anywhere", "N/A", "N/A",
                   "not represented on any frontend" if usd > 0 else "no cost"]
        else:
            clm_t += claimed
            diff = usd - claimed
            status = ("understated on site" if diff > 0.005
                      else "overstated on site" if diff < -0.005 else "matches")
            row = [app, what, usd, where, claimed, round(diff, 2), status]
        ws.append(row)
        r = ws.max_row
        if claimed is None and usd > 0:
            for c in ws[r]:
                c.fill = NA
        elif claimed is not None and usd - claimed > 0.005:
            ws.cell(r, 6).fill = OVER
            ws.cell(r, 6).font = BOLD
    ws.append([])
    ws.append(["TOTAL", "", round(inv_t, 2), "", round(clm_t, 2) if clm_t else "N/A",
               round(inv_t - clm_t, 2) if clm_t else "N/A", ""])
    for c in ws[ws.max_row]:
        c.font, c.fill = BOLD, TOTROW
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.border = THIN
            if c.column in (3, 5, 6) and isinstance(c.value, (int, float)):
                c.number_format = '"$"#,##0.00'
    autosize(ws, [34, 52, 15, 42, 18, 12, 30])
    ws.freeze_panes = "A2"
    return inv_t, clm_t


def main():
    data = load()
    wb = Workbook()
    wb.remove(wb.active)

    # ---------- summary ----------
    ws = wb.create_sheet("Summary")
    ws.append(["SLM project — Modal spend vs what the websites claim"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Account", "Invoiced (USD)", "Apps", "Largest single item"])
    style_header(ws, 3)
    grand = 0.0
    for a in ACCOUNTS:
        apps = data[a]["apps"]
        t = sum(x["usd"] for x in apps)
        grand += t
        top = max(apps, key=lambda x: x["usd"])
        ws.append([a, round(t, 2), len(apps), f'{top["app"]}  (${top["usd"]:.2f})'])
    ws.append(["TOTAL INVOICED", round(grand, 2), "", ""])
    for c in ws[ws.max_row]:
        c.font, c.fill = BOLD, TOTROW
    ws.append([])
    ws.append(["Stated budget", 120.00, "", "user-reported combined spend"])
    ws.append(["Unattributed", round(120 - grand, 2), "",
               "not in the Ephemeral App Breakdown — likely storage/volume, "
               "non-ephemeral apps, or credit-vs-charge rounding"])
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for c in row:
            if c.column == 2 and isinstance(c.value, (int, float)):
                c.number_format = '"$"#,##0.00'
    autosize(ws, [22, 18, 10, 60])

    # ---------- spend by phase ----------
    ws = wb.create_sheet("Spend by phase")
    ws.append(["Phase", "Invoiced (USD)", "% of invoiced", "Shown on frontends?"])
    style_header(ws)
    buckets: dict[str, float] = {}
    for a in ACCOUNTS:
        for x in data[a]["apps"]:
            buckets[PHASE.get(x["app"], "Builds / exports")] = \
                buckets.get(PHASE.get(x["app"], "Builds / exports"), 0) + x["usd"]
    shown = {
        "125M pretraining": "yes — but understated (see Reconciliation)",
        "Fine-tuning (SFT/RAFT)": "yes — understated",
        "Alignment (DPO/RLAIF)": "yes — understated",
        "Evaluation": "NO — never shown on any site",
        "Earlier SFT study": "NO — different project",
        "Probes & tests": "NO",
        "Builds / exports": "NO",
    }
    for k, v in sorted(buckets.items(), key=lambda kv: -kv[1]):
        ws.append([k, round(v, 2), round(v / grand * 100, 1), shown.get(k, "NO")])
    ws.append(["TOTAL", round(grand, 2), 100.0, ""])
    for c in ws[ws.max_row]:
        c.font, c.fill = BOLD, TOTROW
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.border = THIN
            if c.column == 2 and isinstance(c.value, (int, float)):
                c.number_format = '"$"#,##0.00'
            if c.column == 3 and isinstance(c.value, (int, float)):
                c.number_format = '0.0"%"'
    autosize(ws, [26, 16, 14, 44])

    # ---------- per-account sheets ----------
    for a in ACCOUNTS:
        account_sheet(wb, a, data[a])

    # ---------- reconciliation ----------
    ws = wb.create_sheet("Reconciliation")
    ws.append(["Work item", "Invoiced (USD)", "Frontend shows (USD)", "Difference",
               "Understated by", "Where it appears"])
    style_header(ws)
    recon = [
        ("125M pretraining — v1", 11.54, 10.71, "e4 site"),
        ("125M pretraining — extension", 2.88, 2.70, "e4 site"),
        ("125M pretraining — e2", 27.47, 26.31, "e4 site"),
        ("125M pretraining — e4", 28.25, 18.06, "e4 site"),
        ("125M pretraining — TOTAL", 70.14, 57.79, "e4 site + all four 125M model sites"),
        ("Gemma SFT+RAFT (both accounts)", 16.66, 4.56, "Gemma QA + RAFT sites"),
        ("Gemma DPO", 0.43, 0.09, "Gemma DPO site"),
        ("Gemma RLAIF (PPO)", 4.23, 1.26, "Gemma RLAIF site"),
        ("500M SFT+RAFT", 2.67, 1.23, "500M sites"),
        ("500M DPO", 0.08, 0.05, "500M DPO site"),
        ("500M RLAIF (PPO)", 0.33, 0.28, "500M RLAIF site"),
        ("125M SFT+RAFT", 0.61, 0.39, "125M QA + RAFT sites"),
        ("125M DPO", 0.02, 0.01, "125M DPO site"),
        ("125M RLAIF (PPO)", 0.08, 0.07, "125M RLAIF site"),
        ("Reward model", 0.06, 0.03, "RLAIF sites"),
    ]
    for name, inv, clm, where in recon:
        d = round(inv - clm, 2)
        ws.append([name, inv, clm, d, f"{inv/clm:.1f}x" if clm else "—", where])
        if d > 0.005:
            ws.cell(ws.max_row, 4).fill = OVER
            ws.cell(ws.max_row, 4).font = BOLD
        if "TOTAL" in name:
            for c in ws[ws.max_row]:
                c.font = BOLD
    ws.append([])
    ws.append(["Evaluation (set1 + set2)", 19.19, "N/A", "N/A", "—",
               "NOT shown on any frontend — the single largest fine-tuning-side cost"])
    for c in ws[ws.max_row]:
        c.fill = NA
        c.font = BOLD
    ws.append(["Earlier 125M SFT study", 1.66, "N/A", "N/A", "—", "separate project, not on these sites"])
    ws.append(["Phase-0 capability probe", 0.18, "N/A", "N/A", "—", "not on these sites"])
    ws.append(["Image builds / exports / smoke tests", 0.04, "N/A", "N/A", "—", "not on these sites"])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.border = THIN
            if c.column in (2, 3, 4) and isinstance(c.value, (int, float)):
                c.number_format = '"$"#,##0.00'
    autosize(ws, [38, 16, 22, 14, 16, 56])
    ws.freeze_panes = "A2"

    out = os.path.join(HERE, "SLM-cost-audit.xlsx")
    wb.save(out)
    print("wrote", out)
    print(f"grand total invoiced: ${grand:.2f}")
    return grand, buckets


if __name__ == "__main__":
    main()
